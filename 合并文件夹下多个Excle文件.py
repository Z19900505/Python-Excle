# -*- coding: UTF-8 -*-

import os, sys, time
count = 3 #最大检测3次，即第一次检测不存在则安装，会有安装失败的情况，就会再来一次
try:
    import xlrd #用于一次读取Excel中的整行数据
    import xlwt
    from openpyxl import load_workbook #用于写入数据
    from tqdm import tqdm, trange #用于显示进度条
except:
    print("import moudle failed, try to install by pip!")
    os.system('pip install xlrd xlwt openpyxl tqdm')

# 无聊之作，打印12星座
for i in range(12):
    print(str(chr(9800+i)),'\t',end=" ")

#0. 获取文件所在文件夹路径
file_path = input("请输入文件夹绝对路径，例如：D:\数据\Data \t")
if not os.path.exists(file_path):
    file_path = input("输入路径不存在，请重新输入正确路径：\t")

#1.获取路径下所有文件，并存入列表
pathss = []   # 存储文件夹内所有文件的路径（包括子目录内的文件）
            # os.walk() 方法用于通过在目录树中游走输出在目录中的文件名，向上或者向下。
            # os.walk() 方法是一个简单易用的文件、目录遍历器，可以帮助我们高效的处理文件、目录方面的事情。
for root, dirs, files in os.walk(file_path):
    path = [os.path.join(root, name) for name in files]
    pathss.extend(path)
print("已读取需要合并的文件列表，共计：%i" %(len(pathss))+ "个文件\n")
print("开始处理...\n")

#2.只提取出需要的Excel文件的路径
files_for_merge=[]
for i in pathss:
    if 'Tracker-sub' in i: # 仅选取文件名包含“Tracker-sub”的文件
        files_for_merge.append(i)

#3.读取各个Excel中的数据，并存入列表
data = []
for i in files_for_merge:
    wb = xlrd.open_workbook(i) #按相应路径列表pathss顺序读取工作簿
    ws = wb.sheet_by_index(0) #选取工作表
    for j in range(10,ws.nrows):  #示例文件数据从第10行开始
        data.append(ws.row_values(j)) #读取整行数据，并存入列表

#4.汇总数据到主Excel文件 
wb_main = load_workbook(file_path + "/FM Reduction Activities Tracker-main.xlsx") #打开需要写入数据的工作簿

try: # 判读当前工作簿workbook是否存在名为'File_List'的工作表sheet
    ws_list = wb_main['File_List']
except:
    wb_main.create_sheet(title= 'File_List', index= 0)
    ws_list = wb_main['File_List']

for row in range(1, len(pathss)+1):
    ws_list.cell(row= row, column= 1, value= pathss[row-1])

ws_main = wb_main['Raw Findings'] #选定需要写入数据的工作表 
for row in range(3, len(data)+3):  # 前两行为表头，从第3行开始
    for col in range(1,18): # 每行数据18列
        ws_main.cell(row=row, column=col, value=data[row-3][col-1]) #写入数据
        
wb_main.save(file_path+"/FM Reduction Activities Tracker-main.xlsx") #保存数据 

for i in tqdm(range(100)):  #显示进度条等同 for i in trange(100)
    time.sleep(0.05)
    pass
print("程序执行完成！")


'''
*********************************************************************************************************************
#1 知识点 append() 与 extend()的区别，都仅接受一个参数
>>> myList= [1, 2.0, 'a']
>>> myList.append(['APP', 123])
>>> myList
[1, 2.0, 'a', ['APP', 123]]
>>> myList.extend([123, 'abc'])
>>> myList
[1, 2.0, 'a', 'APP', 123, 'abc']
*********************************************************************************************************************
#2 知识点 os.walk()
os.walk可遍历一个目录内各个子目录和子文件。它先遍历当前目录，返回三个值，分别是目录的路径，目录下子目录的名字，文件的名字。
再遍历子目录，同样返回子目录的路径，子目录下的子目录的名字，子目录内的文件的名字。若还有子目录，则继续遍历，直到所有目录被遍历。
因此需要三个变量root, dirs, files去接收它的返回值。由于path本身是一个列表，需要将每次采集到的文件汇总到一个列表中去，我们使用extend来完成。
*********************************************************************************************************************
#3 知识点 进度条(tqdm库)
from tqdm import tqdm,trange
import time

# 于可以迭代的对象都可以使用下面这种方式，来实现可视化进度，非常方便
for i in tqdm(range(100)):
    time.sleep(0.1)
    pass

# 在使用tqdm的时候，可以将tqdm(range(100))替换为trange(100)
for i in trange(100):
    time.sleep(0.1)
    pass

# 通过tqdm提供的set_description方法可以实时查看每次处理的数据
pbar = tqdm(["a","b","c","d"])
for c in pbar:
    time.sleep(1)
    pbar.set_description("Processing %s"%c)

# 通过update方法可以控制每次进度条更新的进度
#total参数设置进度条的总长度
with tqdm(total=100) as pbar:
    for i in range(100):
        time.sleep(0.05)
        #每次更新进度条的长度
        pbar.update(1)

# 除了使用with之外，还可以使用另外一种方法实现上面的效果
#total参数设置进度条的总长度
pbar = tqdm(total=100)
for i in range(100):
    time.sleep(0.05)
    #每次更新进度条的长度
    pbar.update(1)
#关闭占用的资源
pbar.close()
*********************************************************************************************************************
'''
