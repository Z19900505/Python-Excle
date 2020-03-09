'''
 # @ Author: Zeng Hugh
 # @ Create Time: 2020-03-09 13:56:50
 # @ Modified by: Zeng Hugh
 # @ Modified time: 2020-03-09 18:11:42
 # @ Description:
 '''
# -*- coding: UTF-8 -*-
import os, sys, time
try:
    from tqdm import tqdm
    from openpyxl import load_workbook
    from IPython.core.interactiveshell import InteractiveShell
except:
    os.system('pip install openpyxl IPython tqdm')

InteractiveShell.ast_node_interactivity = 'all'

class Auto_Bulk_Lable(object):
    def __init__(self, info_folder_path, lable_sheet_path):
        self.info_folder_path = info_folder_path  #  待写入标签信息文档目录
        self.lable_sheet_path = lable_sheet_path  #  标签信息将要写入的工作表路径及文件名
        
    # 获取物料信息表文件完整路径，建议绝对路径避免出错
    def read_info_write_lable(self):
        if not os.path.exists(self.info_folder_path):
            print("输入文件夹路径不存在，请重新输入!!!")
            exit()
        else:
            lable_info_pathss = []
            for root, dirs, files in os.walk(self.info_folder_path):
                path = [os.path.join(root, name) for name in files]
                lable_info_pathss.extend(path)
        # 输出目录下文件列表, 包括子目录
        for i in range(len(lable_info_pathss)):
            print(lable_info_pathss[i])
        print("已读取指定目录下文件列表，共计：%i" %(len(lable_info_pathss))+ "个文件。")
        print("开始处理...")
        # 于可以迭代的对象都可以使用下面这种方式，来实现可视化进度，非常方便
        for i in tqdm(range(100)):
            time.sleep(0.05)
            pass
        # 从目录下待写入标签信息工作表读取并处理信息，存入列表
        files_for_write = []
        for i in lable_info_pathss:
            # if "lable_info" in i: #  根据指定条件筛选出需要处理的文件名
            files_for_write.append(i)
        lable_info=[]  # 用于存储需要后续写入的标签信息
        for i in files_for_write:
            info_workbook = load_workbook(i)  #按相应路径列表pathss顺序读取工作簿
            info_worksheet = info_workbook['info']
            for row in range(2, info_worksheet.max_row+1):
                brand = info_worksheet['A' + str(row)].value
                if brand:   # 判断“品牌”是否为空值，若为空值则不进行小写字母的转换
                            # 若无这行，当遇到“品牌”为空值时，程序会报错，因为空值无法进行大小写转换
                    brand=brand.upper() #将“品牌”中的小写字母全部转换成大写字母
                typ = info_worksheet['B' + str(row)].value # 获取单元格中的数据
                pn = info_worksheet['C' + str(row)].value
                lotno = info_worksheet['D' + str(row)].value
                date = info_worksheet['E' + str(row)].value
                if date:    # 判断“日期时间”是否为空值，若为空值则不进行日期获取
                            # 若无这行，当遇到“日期时间”为空值时，程序会报错，因为空值无法进行日期获取
                    date = date.date() #只获取日期时间中的日期，比如2019-2-20，不需要具体时间
                    # date = time.strftime("%Y-%m-%d", time.localtime()) 
                quantity = str(info_worksheet['F' + str(row)].value)+" pcs" # 在数量后面加上“pcs”字样
                data = {
                    "brand":brand,
                    "typ":typ,
                    "pn":pn,
                    "lotno":lotno,
                    "date":date,
                    "quantity":quantity
                    }
                lable_info.append(data)
        '''
        # 逐条显示已读取的标签信息内容
        for i in range(len(lable_info)): 
            print(lable_info[i])
        '''
        # 开始写入获取的标签信息
        if not os.path.exists(self.lable_sheet_path):
            print("待写入标签文件不存在，请重新输入!!!")
        else:
            lable_workbook = load_workbook(self.lable_sheet_path)
            lable_worksheet = lable_workbook['lable']
            k = 0
            for i in range(2, 9, 3):   # 列遍历，物料表签2列，间隔1列，共3列。A4纸每页横行放3个标签
                for j in range(1, round(len(lable_info)*7/3), 7):  # 行遍历，物料表签6行，加间隔1行，共7行
                    if k < len(lable_info):  #当数据条数不是3的整数倍
                        lable_worksheet.cell(row=j,   column=i-1).value = "品牌Brand"
                        lable_worksheet.cell(row=j,   column=i).value = lable_info[k]['brand']
                        lable_worksheet.cell(row=j+1, column=i-1).value = "型号Type"
                        lable_worksheet.cell(row=j+1, column=i).value = lable_info[k]['typ']
                        lable_worksheet.cell(row=j+2, column=i-1).value = "物料编号Item P/N"
                        lable_worksheet.cell(row=j+2, column=i).value = lable_info[k]['pn']
                        lable_worksheet.cell(row=j+3, column=i-1).value = "生产批号Lot No."
                        lable_worksheet.cell(row=j+3, column=i).value = lable_info[k]['lotno']
                        lable_worksheet.cell(row=j+4, column=i-1).value = "数量Quantity"
                        lable_worksheet.cell(row=j+4, column=i).value = lable_info[k]['quantity']
                    k += 1
            lable_workbook.save(self.lable_sheet_path)

if __name__=='__main__':
    # 无聊之作，打印12星座
    for i in range(12):
        print(str(chr(9800+i)),'\t',end=" ")
    info_folder_path = input("\n请输入待写入的标签信息文件目录的路径：")
    lable_sheet_path = input("\n请输入需要写入的标签工作簿路径及名称：")
    Auto_Bulk_Lable(info_folder_path, lable_sheet_path).read_info_write_lable()
    for i in range(12):
        print(str(chr(9800+i)),'\t',end=" ")
    print('\n完成啦！*^_^*')
